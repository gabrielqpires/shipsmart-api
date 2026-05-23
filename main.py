import io
import csv
import warnings
from datetime import date, timedelta, datetime

from fastapi import FastAPI, File, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
app = FastAPI(title="Shipsmart Fluxo de Caixa API")

@app.get("/")
def health():
    return {"status": "ok", "service": "Shipsmart Fluxo de Caixa"}

# ─── Helpers ──────────────────────────────────────────────────────────────────
def pbr(s):
    if not s or s.strip() in ('', ',00', '-'): return 0.0
    try: return float(s.strip().replace('.','').replace(',','.'))
    except: return 0.0

def pdate(s):
    if not s or not s.strip(): return None
    try: return datetime.strptime(s.strip(), '%Y-%m-%d').date()
    except:
        try: return datetime.strptime(s.strip(), '%d/%m/%Y').date()
        except: return None

def fill(h): return PatternFill('solid', start_color=h, fgColor=h)
def fnt(bold=False, sz=10, color='000000', name='Arial'):
    return Font(name=name, bold=bold, size=sz, color=color)
def aln(h='left', v='center', wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

T   = Side(style='thin', color='8496B0')
T2  = Side(style='thin', color='FFFFFF')
BDR  = Border(left=T,  right=T,  top=T,  bottom=T)
BDR2 = Border(left=T2, right=T2, top=T2, bottom=T2)
BRL      = '#,##0.00'
DASH_FMT = '#,##0.00;-#,##0.00;"-"'
DATE_FMT = 'DD/MM/YYYY'

AZUL='0D1B5E'; AZUL2='1A2E7C'; AMARELO='F5C842'; BRANCO='FFFFFF'
CINZA_LT='F0F2F8'; CINZA_MED='D8DCE8'; VERDE_LT='E2EFDA'; VERDE_TX='375623'
VERM_LT='FDECEA'; VERM_TX='9C0006'; AMARELO_LT='FFF8DC'; LARANJA_LT='FCE4D6'
THIN_G = Side(style='thin', color='D0D0D0')
BDR_G  = Border(left=THIN_G, right=THIN_G, top=THIN_G, bottom=THIN_G)

# ─── Parse CSV ────────────────────────────────────────────────────────────────
def parse_csv(content: bytes) -> list[dict]:
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text), delimiter=';')
    return [dict(row) for row in reader]

# ─── Gerar XLSX ───────────────────────────────────────────────────────────────
def gerar_xlsx(rows_p: list[dict], rows_r: list[dict]) -> bytes:
    import os
    from datetime import timezone, timedelta as td
    # Força fuso de Brasília (UTC-3)
    HOJE = (datetime.now(timezone(td(hours=-3)))).date()
    FIM  = HOJE + timedelta(days=30)
    datas = [HOJE + timedelta(days=i) for i in range(31)]

    dia_semana = HOJE.weekday()
    seg_semana = HOJE - timedelta(days=dia_semana)
    dom_semana = seg_semana + timedelta(days=6)

    # ── Agregar receber ───────────────────────────────────────────────────────
    r_atras = 0.0
    r_diario = {}
    larr_diario = {}

    for r in rows_r:
        v = pbr(r.get('Valor a Receber',''))
        if v <= 0: continue
        prev = pdate(r.get('Previsão de Recebimento',''))
        if not prev: continue
        is_larr = 'LARROUDE' in (r.get('Cliente (Nome Fantasia)','') or '').upper()
        if prev < HOJE:
            r_atras += v
        else:
            r_diario[prev] = r_diario.get(prev, 0.0) + v
            if is_larr:
                larr_diario[prev] = larr_diario.get(prev, 0.0) + v

    # Média diária — recebimentos no prazo último mês ÷ 30
    MES_ATRAS = HOJE - timedelta(days=30)
    total_no_prazo = 0.0
    for r in rows_r:
        v = pbr(r.get('Valor Recebido',''))
        if v <= 0: continue
        prev  = pdate(r.get('Previsão de Recebimento',''))
        ult   = pdate(r.get('Último Recebimento',''))
        if not prev or not ult: continue
        if prev == ult and MES_ATRAS <= ult <= HOJE:
            total_no_prazo += v
    media_diaria = total_no_prazo / 30

    # ── Agregar pagar ─────────────────────────────────────────────────────────
    p_atras = 0.0
    p_diario = {}
    for r in rows_p:
        v = pbr(r.get('Valor a Pagar',''))
        if v <= 0: continue
        prev = pdate(r.get('Previsão de Pagamento',''))
        if not prev: continue
        if prev < HOJE:
            p_atras += v
        else:
            p_diario[prev] = p_diario.get(prev, 0.0) + v

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook(); wb.remove(wb.active)
    COL_ATRAS=4; COL_D0=5; N=len(datas); LAST_COL=COL_D0+N-1

    # ══════════ FLUXO_DIARIO ══════════════════════════════════════════════════
    ws = wb.create_sheet('FLUXO_DIARIO')
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85

    for row in range(1,10):
        ws.row_dimensions[row].height = 14 if row!=3 else 22
        for col in range(1, LAST_COL+3):
            ws.cell(row=row,column=col).fill = fill(AZUL)
    ws.row_dimensions[8].height = 6

    ws.merge_cells('A1:G9')
    tc=ws.cell(row=1,column=1)
    tc.value=f'FLUXO DE CAIXA DIÁRIO  |  {HOJE.strftime("%d/%m/%Y")}'
    tc.font=fnt(bold=True,sz=12,color=BRANCO); tc.fill=fill(AZUL)
    tc.alignment=aln('left','center')

    ws.merge_cells('H1:I9')
    logo=ws.cell(row=1,column=8)
    logo.value='SHiP\n━━━\nSMART'
    logo.font=Font(name='Arial',bold=True,size=16,color=AMARELO)
    logo.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    logo.fill=fill(AZUL)

    ws.row_dimensions[10].height=22
    for col in range(1,LAST_COL+3):
        c=ws.cell(row=10,column=col)
        c.fill=fill(AZUL2); c.border=BDR2
        c.font=fnt(bold=True,color=BRANCO,sz=9); c.alignment=aln('center')

    ws.cell(row=10,column=COL_ATRAS).value='Atrasados'
    ws.cell(row=10,column=COL_ATRAS).fill=fill('142060')
    ws.cell(row=10,column=COL_ATRAS).font=fnt(bold=True,color=AMARELO,sz=9)

    for i,d in enumerate(datas):
        col=COL_D0+i
        c=ws.cell(row=10,column=col)
        c.value=d; c.number_format=DATE_FMT
        if i==0: c.fill=fill(AMARELO); c.font=fnt(bold=True,color=AZUL,sz=9)
        else: c.fill=fill(AZUL2); c.font=fnt(bold=True,color=BRANCO,sz=9)

    LINHAS={
        11:('Média diária receb (previsão)',False,CINZA_LT,'555555',18),
        12:('A receber',True,CINZA_LT,'000000',18),
        13:('Previsão recebimento Larroude',False,AMARELO_LT,'000000',18),
        14:('Previsão janela',False,AMARELO_LT,'000000',18),
        15:('Previsão não receber Larroude',False,CINZA_LT,VERM_TX,18),
        16:('Total Geral',True,AZUL2,BRANCO,20),
        17:None,
        18:('(-) Total despesas',True,CINZA_LT,VERM_TX,18),
        19:None,
        21:('Entradas',False,CINZA_LT,VERDE_TX,18),
        22:('Saída',False,CINZA_LT,VERM_TX,18),
        23:('Saldo Final',True,AZUL,BRANCO,22),
    }

    for row,info in LINHAS.items():
        if info is None:
            ws.row_dimensions[row].height=5
            for col in range(1,LAST_COL+3): ws.cell(row=row,column=col).fill=fill('E8EBF5')
            continue
        label,bold,bg,fc,h=info
        ws.row_dimensions[row].height=h
        ws.merge_cells(f'A{row}:C{row}')
        c=ws.cell(row=row,column=1)
        c.value=label; c.font=fnt(bold=bold,sz=10,color=fc)
        c.fill=fill(bg); c.alignment=aln(indent=1); c.border=BDR
        for col in [2,3]:
            ws.cell(row=row,column=col).fill=fill(bg)
            ws.cell(row=row,column=col).border=BDR
        ws.cell(row=row,column=COL_ATRAS).fill=fill(bg)
        ws.cell(row=row,column=COL_ATRAS).border=BDR

    ws.row_dimensions[20].height=22
    for col in range(1,LAST_COL+3):
        c=ws.cell(row=20,column=col)
        c.fill=fill(AZUL2); c.border=BDR2
        c.font=fnt(bold=True,color=BRANCO,sz=9); c.alignment=aln('center')
    ws.cell(row=20,column=COL_ATRAS).value='Saldo'
    ws.cell(row=20,column=COL_ATRAS).font=fnt(bold=True,color=AMARELO,sz=9)
    for i,d in enumerate(datas):
        col=COL_D0+i; c=ws.cell(row=20,column=col)
        c.value=d; c.number_format=DATE_FMT
        if i==0: c.fill=fill(AMARELO); c.font=fnt(bold=True,color=AZUL,sz=9)
        else: c.fill=fill(AZUL2); c.font=fnt(bold=True,color=BRANCO,sz=9)

    def sv(ws,row,col,val,bg,fc='000000',bold=False):
        c=ws.cell(row=row,column=col); c.value=val
        c.number_format=DASH_FMT if val==0 else BRL
        c.font=fnt(bold=bold,sz=9,color=fc if val!=0 else 'AAAAAA')
        c.fill=fill(bg); c.border=BDR; c.alignment=aln('right')

    def sa(ws,row,val,bg,fc='000000',bold=False):
        c=ws.cell(row=row,column=COL_ATRAS); c.value=val
        c.number_format=DASH_FMT if val==0 else BRL
        c.font=fnt(bold=bold,sz=9,color=fc if val!=0 else 'AAAAAA')
        c.fill=fill(bg); c.border=BDR; c.alignment=aln('right')

    sa(ws,11,0,CINZA_LT)
    for i,d in enumerate(datas):
        sv(ws,11,COL_D0+i, media_diaria if d.weekday()<5 else 0, CINZA_LT,'555555')

    sa(ws,12,r_atras,CINZA_LT,fc=VERM_TX,bold=True)
    ws.cell(row=12,column=COL_ATRAS).font=fnt(bold=True,sz=9,color=VERM_TX)
    for i,d in enumerate(datas):
        sv(ws,12,COL_D0+i,r_diario.get(d,0.0),CINZA_LT)

    for row in [13,14]:
        sa(ws,row,0,AMARELO_LT)
        for i in range(N):
            c=ws.cell(row=row,column=COL_D0+i)
            c.fill=fill(AMARELO_LT); c.border=BDR
            c.number_format=BRL; c.alignment=aln('right')

    sa(ws,15,0,CINZA_LT)
    for i,d in enumerate(datas):
        v=larr_diario.get(d,0.0)
        c=ws.cell(row=15,column=COL_D0+i)
        c.value=v; c.number_format=DASH_FMT if v==0 else BRL
        c.font=fnt(sz=9,color=VERM_TX if v>0 else 'AAAAAA',bold=(v>0))
        c.fill=fill(CINZA_LT); c.border=BDR; c.alignment=aln('right')

    sa(ws,16,r_atras,AZUL2,fc=BRANCO,bold=True)
    ws.cell(row=16,column=COL_ATRAS).font=fnt(bold=True,sz=9,color=BRANCO)
    for i in range(N):
        col=COL_D0+i; cl=get_column_letter(col)
        c=ws.cell(row=16,column=col)
        c.value=f'=IFERROR({cl}12,0)+IFERROR({cl}13,0)+IFERROR({cl}14,0)-IFERROR({cl}15,0)'
        c.number_format=BRL; c.font=fnt(bold=True,sz=9,color=BRANCO)
        c.fill=fill(AZUL2); c.border=BDR; c.alignment=aln('right')

    sa(ws,18,p_atras,CINZA_LT,fc=VERM_TX,bold=True)
    ws.cell(row=18,column=COL_ATRAS).font=fnt(bold=True,sz=9,color=VERM_TX)
    for i,d in enumerate(datas):
        v=p_diario.get(d,0.0)
        bg=VERM_LT if v>0 else CINZA_LT; fc2=VERM_TX if v>0 else 'AAAAAA'
        sv(ws,18,COL_D0+i,v,bg,fc2,bold=(v>0))

    sa(ws,21,r_atras,VERDE_LT,fc=VERDE_TX)
    sa(ws,22,p_atras,VERM_LT,fc=VERM_TX)
    sa(ws,23,r_atras-p_atras,AZUL,fc=BRANCO,bold=True)
    ws.cell(row=23,column=COL_ATRAS).font=fnt(bold=True,sz=9,color=BRANCO)

    for i in range(N):
        col=COL_D0+i; cl=get_column_letter(col); pcl=get_column_letter(col-1)
        for row,formula,bg,fc2 in [
            (21,f'=IFERROR({cl}16,0)',VERDE_LT,VERDE_TX),
            (22,f'=IFERROR({cl}18,0)',VERM_LT,VERM_TX),
        ]:
            c=ws.cell(row=row,column=col)
            c.value=formula; c.number_format=BRL
            c.font=fnt(sz=9,color=fc2); c.fill=fill(bg); c.border=BDR; c.alignment=aln('right')
        c23=ws.cell(row=23,column=col)
        c23.value=f'=IFERROR(B30,0)+IFERROR({cl}21,0)-IFERROR({cl}22,0)' if i==0 \
                  else f'=IFERROR({pcl}23,0)+IFERROR({cl}21,0)-IFERROR({cl}22,0)'
        c23.number_format=BRL; c23.font=fnt(bold=True,sz=9,color=BRANCO)
        c23.fill=fill(AZUL); c23.border=BDR; c23.alignment=aln('right')

    ws.row_dimensions[25].height=14
    ws.merge_cells('A26:C26')
    th=ws.cell(row=26,column=1,value='💰  SALDO DOS BANCOS  (preencher semanalmente)')
    th.font=fnt(bold=True,sz=10,color=BRANCO); th.fill=fill(AZUL2)
    th.alignment=aln('center'); th.border=BDR
    ws.cell(row=26,column=2).fill=fill(AZUL2); ws.cell(row=26,column=2).border=BDR
    ws.cell(row=26,column=3).fill=fill(AZUL2); ws.cell(row=26,column=3).border=BDR
    ws.row_dimensions[26].height=22

    for ci,h in enumerate(['Banco / Conta','Saldo (R$)','Atualizado em'],1):
        c=ws.cell(row=27,column=ci,value=h)
        c.font=fnt(bold=True,sz=9,color=AZUL); c.fill=fill(CINZA_MED)
        c.border=BDR; c.alignment=aln('center')
    ws.row_dimensions[27].height=18

    for i,conta in enumerate(['Itaú Unibanco','Mercado Pago','Omie.CASH','Outros']):
        row=28+i; bg=BRANCO if i%2==0 else CINZA_LT
        ws.cell(row=row,column=1,value=conta).font=fnt(sz=9)
        ws.cell(row=row,column=1).fill=fill(bg); ws.cell(row=row,column=1).border=BDR
        cb=ws.cell(row=row,column=2,value=0)
        cb.fill=fill(AMARELO_LT); cb.border=BDR; cb.number_format=BRL
        cb.alignment=aln('right'); cb.font=fnt(sz=9)
        cc=ws.cell(row=row,column=3)
        cc.fill=fill(CINZA_LT); cc.border=BDR
        cc.number_format=DATE_FMT; cc.alignment=aln('center'); cc.font=fnt(sz=9,color='888888')
        ws.row_dimensions[row].height=18

    ws.cell(row=32,column=1,value='TOTAL').font=fnt(bold=True,sz=9,color=BRANCO)
    ws.cell(row=32,column=1).fill=fill(AZUL); ws.cell(row=32,column=1).border=BDR
    ws.cell(row=32,column=2).value='=SUM(B28:B31)'
    ws.cell(row=32,column=2).font=fnt(bold=True,sz=10,color=AMARELO)
    ws.cell(row=32,column=2).fill=fill(AZUL); ws.cell(row=32,column=2).border=BDR
    ws.cell(row=32,column=2).number_format=BRL; ws.cell(row=32,column=2).alignment=aln('right')
    ws.cell(row=32,column=3).fill=fill(AZUL); ws.cell(row=32,column=3).border=BDR
    ws.row_dimensions[32].height=20

    nota=ws.cell(row=33,column=1)
    nota.value=f'⚑  O Saldo Final usa o total acima como saldo inicial de {HOJE.strftime("%d/%m/%Y")}'
    nota.font=fnt(sz=8,color='888888'); ws.merge_cells('A33:F33')

    ws.column_dimensions['A'].width=14; ws.column_dimensions['B'].width=14
    ws.column_dimensions['C'].width=14
    ws.column_dimensions[get_column_letter(COL_ATRAS)].width=16
    for i in range(N): ws.column_dimensions[get_column_letter(COL_D0+i)].width=14
    ws.freeze_panes=f'{get_column_letter(COL_D0)}11'

    # ══════════ TRANSPORTADORAS ═══════════════════════════════════════════════
    def make_transp(nome_aba, nome_forn):
        wt=wb.create_sheet(nome_aba)
        wt.sheet_view.showGridLines=False

        df_ab=[r for r in rows_p
               if nome_forn.upper() in (r.get('Fornecedor (Nome Fantasia)','') or '').upper()
               and pbr(r.get('Valor a Pagar',''))>0]

        df_atras=[r for r in df_ab if (pdate(r.get('Previsão de Pagamento','')) or date.max) <= HOJE]
        df_avenc =[r for r in df_ab if (pdate(r.get('Previsão de Pagamento','')) or date.min) > HOJE]

        df_atras.sort(key=lambda r: pdate(r.get('Previsão de Pagamento','')) or date.max)
        df_avenc.sort( key=lambda r: pdate(r.get('Previsão de Pagamento','')) or date.max)

        wt.merge_cells('A1:J1')
        t=wt.cell(row=1,column=1)
        t.value=f'CONTAS A PAGAR — {nome_forn.upper()}  |  {HOJE.strftime("%d/%m/%Y")}'
        t.font=fnt(bold=True,sz=13,color=BRANCO); t.fill=fill(AZUL); t.alignment=aln('center')
        wt.row_dimensions[1].height=30

        kpi_labels=[
            ('Vencido',    VERM_LT,   VERM_TX),
            ('A Vencer',   AMARELO_LT,'5C3D00'),
            ('Total Geral',AZUL2,      BRANCO),
        ]
        wt.row_dimensions[2].height=8; wt.row_dimensions[3].height=30; wt.row_dimensions[4].height=18

        for ki,(klbl,kbg,kfc) in enumerate(kpi_labels):
            cs=ki*2+1; ce=cs+1
            for r2 in [3,4]: wt.merge_cells(f'{get_column_letter(cs)}{r2}:{get_column_letter(ce)}{r2}')
            ck=wt.cell(row=3,column=cs,value=0); ck.number_format=BRL
            ck.font=fnt(bold=True,sz=13,color=kfc); ck.fill=fill(kbg)
            ck.alignment=aln('center'); ck.border=BDR
            wt.cell(row=3,column=cs+1).fill=fill(kbg); wt.cell(row=3,column=cs+1).border=BDR
            cl2=wt.cell(row=4,column=cs,value=klbl)
            cl2.font=fnt(sz=9,color=kfc); cl2.fill=fill(kbg); cl2.alignment=aln('center'); cl2.border=BDR
            wt.cell(row=4,column=cs+1).fill=fill(kbg); wt.cell(row=4,column=cs+1).border=BDR

        COLS=['Nota Fiscal','Previsão de Pagamento','Valor da Conta','_CONTESTADO','Valor a Pagar','Observação']
        LABELS=['Nota Fiscal','Previsão Pgto','Valor Fatura','Valor Contestado (segundo planilha [FIN] AP BR Faturas Pago x Contestado)','Valor a Pagar','Observação']
        WIDTHS=[18,16,16,36,16,0]
        for ci,w in enumerate(WIDTHS,1):
            if w>0: wt.column_dimensions[get_column_letter(ci)].width=w
        obs_max=max((len(r.get('Observação','') or '') for r in df_ab), default=20)
        wt.column_dimensions['F'].width=max(20,min(80,obs_max*0.9))

        def write_sec(start,title,rows_sec,title_bg,bg1,bg2):
            if not rows_sec: return start
            wt.merge_cells(f'A{start}:J{start}')
            ts2=wt.cell(row=start,column=1)
            total_sec=sum(pbr(r.get('Valor a Pagar','')) for r in rows_sec)
            ts2.value=f'  {title}  ({len(rows_sec)} fatura{"s" if len(rows_sec)>1 else ""} — R$ {total_sec:,.2f})'
            ts2.font=fnt(bold=True,sz=10,color=BRANCO); ts2.fill=fill(title_bg)
            ts2.alignment=aln(); ts2.border=BDR; wt.row_dimensions[start].height=20

            for ci,(col,lbl) in enumerate(zip(COLS,LABELS),1):
                c=wt.cell(row=start+1,column=ci,value=lbl)
                c.font=fnt(bold=True,sz=9,color=AZUL); c.fill=fill(CINZA_MED)
                c.border=BDR; c.alignment=aln('center')
            wt.row_dimensions[start+1].height=18

            for ri,r2 in enumerate(rows_sec):
                r3=start+2+ri; bg=bg1 if ri%2==0 else bg2
                prev=pdate(r2.get('Previsão de Pagamento',''))
                dias=(HOJE-prev).days if prev else 0
                for ci,col in enumerate(COLS,1):
                    v=r2.get(col,'') or ''
                    c=wt.cell(row=r3,column=ci)
                    c.fill=fill(bg); c.border=BDR; c.font=fnt(sz=9)
                    if col=='Valor a Pagar':
                        cl_fatura    = get_column_letter(COLS.index('Valor da Conta')+1)
                        cl_contestado= get_column_letter(COLS.index('_CONTESTADO')+1)
                        c.value=f'={cl_fatura}{r3}-{cl_contestado}{r3}'
                        c.number_format=BRL
                        c.alignment=aln('right'); c.font=fnt(sz=9,bold=True,color=VERM_TX)
                    elif col=='_CONTESTADO':
                        c.value=pbr(r2.get('Desconto','') or ''); c.number_format=BRL; c.alignment=aln('right')
                    elif col=='Valor da Conta':
                        c.value=pbr(r2.get('Valor da Conta','') or ''); c.number_format=BRL; c.alignment=aln('right')
                    elif col=='Previsão de Pagamento':
                        c.value=str(prev) if prev else ''; c.alignment=aln('center')
                        if dias>60: c.font=fnt(sz=9,color=VERM_TX,bold=True)
                        elif dias>30: c.font=fnt(sz=9,color='CC6600')
                    else:
                        c.value=v
                        if col=='Observação': c.alignment=Alignment(wrap_text=True,vertical='top')
                wt.row_dimensions[r3].height=16
            return start+2+len(rows_sec)

        cur=6; wt.row_dimensions[cur].height=10; cur+=1

        start_atras = cur
        cur=write_sec(cur,'🔴  VENCIDOS',df_atras,'A02020',VERM_LT,'FFD7D7')
        end_atras = cur - 1
        if df_atras: wt.row_dimensions[cur].height=8; cur+=1

        start_avenc = cur
        cur=write_sec(cur,'🟡  A VENCER',df_avenc,'7A6200',AMARELO_LT,'FFF0B0')
        end_avenc = cur - 1

        col_e = get_column_letter(5)
        def kpi_formula(start, end):
            if start > end: return 0
            return f'=IFERROR(SUM({col_e}{start+2}:{col_e}{end}),0)'

        kpi_values = [
            kpi_formula(start_atras, end_atras),
            kpi_formula(start_avenc, end_avenc),
            f'=IFERROR(SUM({col_e}{start_atras+2}:{col_e}{end_avenc}),0)',
        ]

        for ki, kval in enumerate(kpi_values):
            cs = ki*2+1
            c = wt.cell(row=3, column=cs)
            c.value = kval
            c.number_format = BRL

        wt.freeze_panes='A5'

    make_transp('FEDEX','FEDEX')
    make_transp('UPS','UPS DO BRASIL')
    make_transp('DHL','DHL EXPRESS')

    # ══════════ BASES (ocultas) ════════════════════════════════════════════════
    def write_base(nome, rows, val_cols, date_cols):
        ws2=wb.create_sheet(nome)
        if not rows: return
        cols=list(rows[0].keys())
        for ci,col in enumerate(cols,1):
            c=ws2.cell(row=1,column=ci,value=col)
            c.font=fnt(bold=True,color=BRANCO,sz=9); c.fill=fill('1D6F42')
            c.alignment=aln('center',wrap=True); c.border=BDR_G
        ws2.row_dimensions[1].height=28
        for ri,row in enumerate(rows):
            for ci,col in enumerate(cols,1):
                v=row.get(col,'') or ''
                if col in val_cols:
                    cell=ws2.cell(row=ri+2,column=ci,value=pbr(v))
                    cell.number_format=BRL
                elif col in date_cols:
                    cell=ws2.cell(row=ri+2,column=ci,value=v)
                    cell.number_format=DATE_FMT
                else:
                    cell=ws2.cell(row=ri+2,column=ci,value=v)
                cell.font=fnt(sz=8); cell.alignment=aln(v='center'); cell.border=BDR_G
                if (ri+2)%2==0: cell.fill=fill('F5F5F5')
        for ci,col in enumerate(cols,1):
            ws2.column_dimensions[get_column_letter(ci)].width=min(max(len(col)+2,8),28)
        ws2.freeze_panes='A2'
        ws2.sheet_state='hidden'

    val_p=['Valor da Conta','Valor Líquido','Impostos Retidos','Desconto','Juros e Multa','Valor Pago','Valor a Pagar']
    val_r=['Valor da Conta','Valor Líquido','Impostos Retidos','Desconto','Juros e Multa','Valor Recebido','Valor a Receber']
    date_p=['Previsão de Pagamento','Último Pagamento','Vencimento','Data de Emissão','Data de Registro']
    date_r=['Previsão de Recebimento','Último Recebimento','Vencimento','Data de Emissão','Data de Registro']
    write_base('BASE_RECEBER',rows_r,val_r,date_r)
    write_base('BASE_PAGAR',rows_p,val_p,date_p)

    order=['FLUXO_DIARIO','FEDEX','UPS','DHL','BASE_RECEBER','BASE_PAGAR']
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


@app.post('/gerar-fluxo')
async def gerar_fluxo(
    csv_pagar:   UploadFile = File(...),
    csv_receber: UploadFile = File(...),
):
    rows_p = parse_csv(await csv_pagar.read())
    rows_r = parse_csv(await csv_receber.read())
    xlsx_bytes = gerar_xlsx(rows_p, rows_r)
    filename = f'Fluxo_Shipsmart_{date.today().isoformat()}.xlsx'
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )
